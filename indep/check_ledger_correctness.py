"""檢查是否有填寫錯誤
* 欄位"已使用" = 所有當項不同使用總時數
* 所有項數的分數加總是否總和為1, 預假核定時算新的項喔!
"""
import re
from collections import defaultdict

import openpyxl as ox
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

pattern_date3 = '(\d{3})(\d{2})(\d{2})'
pattern_usage = '(\([\d/?]+\))'
pattern_detectfrac = '\(([\d?]+)/([\d?]+)\)'

wb: Workbook = ox.load_workbook('holiday_ledger/組改期間役男榮譽假清冊3_改temp.xlsx')
a2j = 'ABCDEFGHIJ'


def from_liyuen(l, doprint=False):
    skip = True
    for s in l:
        if s.title == '北辦': continue
        if s.title != '張瀚文' and skip:
            continue
        else:
            skip = False
        if doprint: print(s.title)
        yield s


def sanitize(code, func, doprint=False):
    for ws in from_liyuen(wb.worksheets, doprint=doprint):
        ws: Worksheet
        jobs = ws['B']
        names = ws['C']
        last_name = None
        col = ws[code]
        for idx, (job, name, entry) in enumerate(zip(jobs, names, col)):
            if idx == 0: continue
            last_name = last_name if name.value is None else name.value
            if entry.value is None:
                continue
            newval = func('預' + last_name if job.value is None else name.value, entry.value, ws=ws, idx=idx)
            matches = re.findall('\([\d/?]+\)', newval)
            assert all(sum(c == '/' for c in match) <= 1 for match in matches)
            if not (len(entry.value.splitlines()) == len(newval.splitlines()) == len(matches)):
                print('PROBLEMOS, reduction of lines!!')
            entry.value = newval


def verify_count():
    unique_denom = {}
    numrtr_left = {}
    antiparticle_pair = defaultdict(int)

    def correct(name, val, ws, idx, **kwargs):
        showname = name.replace('預', '')
        totused = ws['G'][idx].value
        for l in val.splitlines():
            token, n, d = re.search('([\w\W]+)' + pattern_detectfrac, l).groups()
            if '預' in name:
                antiparticle_pair[token] = 1 | antiparticle_pair[token]
            elif '預' in token:
                antiparticle_pair[token.replace('預', '')] = 2 | antiparticle_pair[token]
            assert d != '?'  # 沒有問號
            assert n != '?'  # 沒有問號
            n, d = map(int, (n, d))
            assert showname, token not in unique_denom or unique_denom[showname, token] == d  # 沒有互不同意的兩種不同分母
            if (showname, token) not in numrtr_left:
                numrtr_left[showname, token] = d - n
            else:
                numrtr_left[showname, token] -= n
            totused -= n
        assert totused == 0  # total of n within a cell should be equal to the column value "已使用"
        return val

    sanitize('I', correct)
    for k, v in numrtr_left.items():
        assert v == 0  # if broken here, fraction doesn't add up to 1
    for k, v in antiparticle_pair.items():
        if not v in [0, 3]:
            print(f"{k} 好像沒有 預假/核銷 都有")


if __name__ == '__main__':
    verify_count()
    # wb.save('holiday_ledger/組改期間役男榮譽假清冊3_改.xlsx')
