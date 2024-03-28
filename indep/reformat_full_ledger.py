"""舊有系統過於髒亂，清理成新的分數格式。
因為是銜接性任務，
沒有重新跑本腳本的必要!
"""
import collections
import math
import operator
import re
import shutil
from collections import defaultdict

import openpyxl as ox
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

pattern_date3 = r'(\d{3})(\d{2})(\d{2})'
pattern_usage = r'(\([\d/?]+\))'
pattern_detectfrac = r'\(([\d?]+)/([\d?]+)\)'

# shutil.copy('holiday_ledger/組改期間役男榮譽假清冊3.xlsx', 'holiday_ledger/組改期間役男榮譽假清冊3_改.xlsx')

target = r'D:\東區督考科業務\1.東區業務\組改期間榮譽假系統\組改期間役男榮譽假清冊自動化.xlsx'  # TODO
wb: Workbook = ox.load_workbook(target)
a2j = 'ABCDEFGHIJ'


def from_liyuen(l, doprint=False):
    skip = True
    for s in l:
        if s.title == '北辦': continue
        if s.title != '張瀚文' and skip:
            continue
        else:
            skip = False
        if doprint: print(s.title)
        yield s


def idxs_cells(num):
    for c in a2j: yield f'{c}{num}'


def sanitize_br():
    for ws in from_liyuen(wb.worksheets):
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    cell.value = cell.value.replace('<br />', '')


def sanitize_header():
    counts = collections.OrderedDict((k, defaultdict(int)) for k in idxs_cells(1))

    for ws in from_liyuen(wb.worksheets):
        # print(ws.title)
        for k in counts.keys():
            counts[k][ws[k].value] += 1

    best = tuple(max(c.items(), key=operator.itemgetter(1)) for k, c in counts.items())
    best = [b[0] for b in best]
    best[-1] = '附註'

    for ws in from_liyuen(wb.worksheets):
        if ws[f'C1'].value != best[2]:
            ws.insert_rows(0)
        for c, v in zip(a2j, best):
            ws[f'{c}1'].value = v


def count_vals(code):
    counts = collections.defaultdict(int)
    for ws in from_liyuen(wb.worksheets):
        for c in ws.columns:
            if c[0].column_letter != code:
                continue
            for e in c[1:]:
                counts[e.value] += 1
    print(counts)


def print_vals(code):
    for ws in from_liyuen(wb.worksheets):
        for c in ws.columns:
            if c[0].column_letter != code:
                continue
            for e in c[1:]:
                if e.value is None: continue
                print(e.value)
                # print(e.value.replace('<br />', ''))


def sanitize(code, func, doprint=False):
    for ws in from_liyuen(wb.worksheets, doprint=doprint):
        ws: Worksheet
        jobs = ws['B']
        names = ws['C']
        last_name = None
        col = ws[code]
        for idx, (job, name, entry) in enumerate(zip(jobs, names, col)):
            if idx == 0: continue
            last_name = last_name if name.value is None else name.value
            if entry.value is None:
                continue
            newval = func('預' + last_name if job.value is None else name.value, entry.value, ws=ws, idx=idx)
            matches = re.findall(r'\([\d/?]+\)', newval)
            assert all(sum(c == '/' for c in match) <= 1 for match in matches)
            if not (len(entry.value.splitlines()) == len(newval.splitlines()) == len(matches)):
                print('PROBLEMOS')
                # func('預' + last_name if job.value is None else name.value, entry.value, ws=ws, idx=idx)
            entry.value = newval


def hasparanthesis(name, v, **kwargs):
    if v.strip() == '':
        return None
    assert '(' in v

    v = v.replace('<br/>', '')
    v = v.replace('</br>', '')
    v = v.replace('</ br>', '')
    v = v.replace('<br />', '')

    v = v.replace('(4)就解', '就解(4)')
    v = v.replace('(4) 就解', '就解(4)')

    v = re.sub(r'\n', '', v)
    if re.search(r'\)(\S)', v):
        v = re.sub(r'\)(\S)', r')\n\g<1>', v)
    if re.search(r'\)(\s+)', v):
        v = re.sub(r'\)(\s+)', r')\n', v)

    v = re.sub(r'\s+(\([\d/]+\))', r'\g<1>', v)
    for _ in range(5):
        v = v.replace('  ', ' ')
        v = v.replace('-', '~')
        v = v.replace(' ~ ', '~')
        v = v.replace(':', '')
    v = v.replace('0800~1700', '全日八')
    v = v.replace('0830~1730', '全日八')
    v = v.replace('下午', '晚四')
    v = v.replace('晚(', '晚四(')
    v = v.replace('早(', '早四(')
    if re.search(pattern_date3, v):
        v = re.sub(pattern_date3, r'\g<1>/\g<2>/\g<3>', v)

    reconstruct = []
    for s in v.splitlines():
        usage = re.search(pattern_usage, s).group(1)
        token = re.sub(r'\([\d/?]+\)', '', s)

        if '/' in usage:  # GOOD already fraction
            reconstruct.append(token + usage)
            continue
        if re.match(r'\*[\S\d\ /~]+', token):  # GOOD 預假核銷
            reconstruct.append(token + usage)
            continue
        if any(keyword in token for keyword in '特 手 自 改 修'.split()):  # GOOD 手動
            reconstruct.append(token + usage)
            continue
        if re.match(r'11\d/\d+/\d+ 全日八', token):  # GOOD
            reconstruct.append(token + usage)
            continue
        if re.fullmatch(r'11\d/\d+/\d+', token):
            reconstruct.append(token + ' 全日八' + usage)
            continue
        if (m := re.match(r'(11\d/\d+/\d+) 0830~1730', token)):
            reconstruct.append(m.group(1) + ' 全日八' + usage)
            continue
        if (m := re.match(r'(11\d/\d+/\d+) 0830~(11\d/\d+/\d+) 1730', token)) and m.group(1) == m.group(2):
            reconstruct.append(m.group(1) + ' 全日八' + usage)
            continue

        if re.match(r'11\d/\d+/\d+~11\d/\d+/\d+ 多日', token):  # GOOD
            reconstruct.append(token + usage)
            continue
        if re.fullmatch(r'11\d/\d+/\d+~11\d/\d+/\d+', token):
            reconstruct.append(token + ' 多日' + usage)
            continue
        if re.match(r'11\d/\d+/\d+ 0830~11\d/\d+/\d+ 1730', token):
            reconstruct.append(token + ' 多日' + usage)
            continue
        if re.match(r'11\d/\d+/\d+ 1330~11\d/\d+/\d+ 1330', token):
            reconstruct.append(token + ' 含半多日' + usage)
            # print(reconstruct[-1])
            continue
        if re.fullmatch(r'11\d/\d+/\d+~\d+/\d+', token):
            reconstruct.append(token + ' 多日' + usage)
            continue
        if re.fullmatch(r'11\d/\d+/\d+~\d+', token):
            reconstruct.append(token + ' 多日' + usage)
            continue

        if (m := re.search(r'[早晚]四', token)):  # GOOD
            reconstruct.append(token + usage)
            continue
        if '晚' in token and '晚四' not in token:
            reconstruct.append(token.replace('晚', '') + '晚四' + usage)
            continue
        if (m := re.match(r'(11\d/\d+/\d+) 0830~1230', token)):
            reconstruct.append(m.group(1) + ' 早四' + usage)
            continue
        if (m := re.match(r'(11\d/\d+/\d+) 0830~1230', token)):
            reconstruct.append(m.group(1) + ' 早四' + usage)
            continue
        if (m := re.match(r'(11\d/\d+/\d+) 1330~1730', token)):
            reconstruct.append(m.group(1) + ' 晚四' + usage)
            continue
        if (m := re.match(r'(11\d/\d+/\d+) (\d{4})~(\d{4}) 當日(\d)', token)):  # GOOD
            reconstruct.append(token + usage)
            continue
        if (m := re.match(r'(11\d/\d+/\d+) (\d{4})~(\d{4})', token)):
            tot = (int(m.group(3)) - int(m.group(2))) // 100
            if int(m.group(2)) < 1245 < int(m.group(3)): tot -= 1  # 午休
            reconstruct.append(token + f' 當日{tot}' + usage)
            continue
        if (m := re.match(r'(11\d/\d+/\d+) (\d{4})~(11\d/\d+/\d+) (\d{4})', token)):
            if m.group(1) == m.group(3):
                tot = (int(m.group(4)) - int(m.group(2))) // 100
                if int(m.group(2)) < 1245 < int(m.group(4)): tot -= 1  # 午休
                if tot >= 8:
                    if tot == 8:
                        raise ValueError
                    reconstruct.append(m.group(1) + ' 全日八' + usage)
                elif tot == 4:
                    if m.group(2) in ['0800', '0830']:
                        reconstruct.append(m.group(1) + ' 早四' + usage)
                    elif m.group(2) in ['1300', '1330']:
                        reconstruct.append(m.group(1) + ' 晚四' + usage)
                    else:
                        reconstruct.append(token + ' 特別' + usage)
                else:
                    reconstruct.append(f'{m.group(1)} {m.group(2)}~{m.group(4)} 當日{tot}' + usage)
            else:
                tot = (int(m.group(4)) - int(m.group(2))) // 100
                if int(m.group(2)) < 1245 < int(m.group(4)): tot -= 1  # 午休
                if tot >= 8:
                    if tot != 8:
                        raise ValueError
                    reconstruct.append(m.group(1) + '~' + m.group(3) + ' 多日' + usage)
                else:
                    reconstruct.append(token + ' 含半多日' + usage)  #
                    # print(reconstruct[-1])
            continue
        reconstruct.append(token + ' 含半多日' + usage)
        # print(reconstruct[-1])

    if '解' in token or '集' in token:
        print('reconstruct: \n' + '\n'.join(reconstruct))

    v = '\n'.join(reconstruct)
    for line in v.splitlines():
        assert len(re.findall(r'11\d', line)) - len(re.findall(r'~', line)) <= 1
    return v


def compress(name, v, **kwargs):
    lines = []
    for line in v.splitlines():
        usage = re.search(pattern_usage, line).group(1)
        token = re.sub(r'\([\d/?]+\)', '', line)
        if '~' in token:
            fr, to = token.split('~')
            fr = fr.replace(' ', '//')
            to = to.replace(' ', '//')
            till = []
            diff = False
            xs = fr.split('/')
            for i, y in enumerate(to.split('/')):
                x = xs[i] if i < len(xs) else ''
                if diff or x != y:
                    diff = True
                    till.append(y)
                if x == y:
                    continue
            to = "/".join(till)
            fr = fr.replace('//', ' ')
            to = to.replace('//', ' ')
            token = f'{fr}~{to}'
            lines.append(token + usage)
        else:
            lines.append(line)
    return '\n'.join(lines)


def precheck_confirm():
    sawtimes = defaultdict(int)
    sawusages = defaultdict(set)

    # format name, token
    def fill_saw(name, v, **kwargs):
        for s in v.splitlines():
            usage = re.search(pattern_usage, s).group(1)
            token = re.sub(pattern_usage, '', s)
            sawtimes[(name, token)] += 1
            sawusages[(name, token)].add(usage)
        return v

    def prefill_fraction(name, v, **kwargs):
        reconstructs = []
        for line in v.splitlines():
            usage = re.search(pattern_usage, line).group(1)
            token = re.sub(pattern_usage, '', line)
            special = True
            if not '/' in usage:
                if '全日八' in token:
                    if usage != '(8)':
                        usage = f'{usage}/(8)'.replace(')/(', '/')
                        special = False
                    else:
                        if sawtimes[(name, token)] == 1:
                            usage = '(8/8)'
                            special = False
                        else:
                            usage = '(?/8)'
                            special = False
                elif '多日' in token or '含半多日' in token:
                    if int(re.search(r'(\d+)', usage).group(1)) <= 8:
                        usage = f'{usage}/(?)'.replace(')/(', '/')
                        special = False
                    elif sawtimes[(name, token)] > 1 and len(sawusages[(name, token)]) == 1:
                        usage = f'(?)/{usage}'.replace(')/(', '/')
                        special = False
                elif '當日' in token:
                    tot = int(re.search(r'當日(\d)', token).group(1))
                    if tot == 1:
                        usage = '(1/1)'
                        special = False
                    elif f'({tot})' != usage:
                        usage = f'{usage}/({tot})'.replace(')/(', '/')
                        special = False
                    elif sawtimes[(name, token)] == 1:
                        usage = f'({tot}/{tot})'
                        special = False
                    else:
                        usage = f'(?/{tot})'
                        special = False
                elif '四' in token:
                    if usage != '(4)':
                        usage = f'{usage}/(4)'.replace(')/(', '/')
                        special = False
                    else:
                        if sawtimes[(name, token)] == 1:
                            usage = '(4/4)'
                            special = False
                        else:
                            usage = '(?/4)'
                            special = False
                elif len(sawusages[(name, token)]) > 1:
                    usage = f'{usage}/(?)'.replace(')/(', '/')
                    special = False
            else:
                special = False
            if special:
                usage = f'{usage}/{usage}'.replace(')/(', '/')
            reconstructs.append(token + usage)
            if special: print(reconstructs[-1])
        v = '\n'.join(reconstructs)
        return v

    def used_override_singleline_and_singletons(name, usage_value, ws, idx):
        used = ws['G'][idx]
        if usage_value is None: return
        if len(usage_value.splitlines()) == 1:
            real_n = used.value
            n, d = re.search(pattern_detectfrac, usage_value).groups()
            if str(real_n) != str(n):
                # print(name, usage_value, end='==>')
                usage_value = re.sub(pattern_detectfrac, f'({real_n}/{d})', usage_value)
                # print(usage_value, end='==>')
                if sawtimes[name, usage_value] == 1:
                    d = real_n
                    usage_value = re.sub(pattern_detectfrac, f'({real_n}/{d})', usage_value)
                    print(usage_value)
                else:
                    pass
                    # print()
        else:
            lines = []
            for l in usage_value.splitlines():
                token, n, d = re.search(r'([\w\W]+)' + pattern_detectfrac, l).groups()
                if sawtimes[name, token] == 1:
                    if n == '?': n = d
                    if d == '?': d = n
                lines.append(
                    re.sub(pattern_detectfrac, f'({n}/{d})', l)
                )
            usage_value = '\n'.join(lines)
        return usage_value

    coverage = {}

    def preemptize(name, val, ws, idx):
        if name.startswith('預'):
            for l in val.splitlines():
                token, n, d = re.search(r'([\w\W]+)' + pattern_detectfrac, l).groups()
                assert (name.replace('預', ''), token) not in coverage
                coverage[(name.replace('預', ''), token)] = n
        else:
            lines = []
            for l in val.splitlines():
                token, n, d = re.search(r'([\w\W]+)' + pattern_detectfrac, l).groups()
                token, usage = re.search(r'([\w\W]+)' + pattern_usage, l).groups()
                assert '/' in usage
                if (name, token) in coverage:
                    lines.append('*預' + token + f'({n}/{coverage[(name, token)]})')
                    # lines.append('*預' + token + usage)
                else:
                    lines.append(l)
            val = '\n'.join(lines)
        return val

    # def fill_prempt(name, val, ws, idx):
    #     used = ws['G'][idx]
    #     if val is None: return
    #     lines = []
    #     for line in val.splitlines():
    #         if '*預' in line:
    #             token, n, d = re.search(r'([\w\W]+)' + pattern_detectfrac)
    #             x = min(used.value,n)
    #             token2 = f'({x}/{n})'
    #         else:
    #             lines.append(line)
    #     val = '\n'.join(lines)
    #     return val

    sanitize('I', fill_saw, doprint=True)
    sanitize('I', prefill_fraction, doprint=True)
    sanitize('I', used_override_singleline_and_singletons, doprint=False)
    sanitize('I', preemptize, doprint=False)
    sawtimes.clear(), sawusages.clear()
    sanitize('I', fill_saw, doprint=False)
    # sanitize('I', fill_prempt, doprint=False)


def iter_solve(name, val, ws, idx, **kwargs):
    tot = ws['G'][idx].value
    ns = []
    lines = []
    for l in val.splitlines():
        token, n, d = re.search(r'([\w\W]+)' + pattern_detectfrac, l).groups()
        ns.append(n)
    if sum(c == '?' for c in ns) == 1:
        for n in ns: tot -= (0 if n == '?' else int(n))
    for l in val.splitlines():
        token, n, d = re.search(r'([\w\W]+)' + pattern_detectfrac, l).groups()
        if n == '?': n = tot
        lines.append(
            re.sub(pattern_detectfrac, f'({n}/{d})', l)
        )
    return '\n'.join(lines)


def verify_count():
    summation_n = defaultdict(int)

    def sum_count_n(name, val, ws, idx, **kwargs):

        for l in val.splitlines():
            token, n, d = re.search(r'([\w\W]+)' + pattern_detectfrac, l).groups()
            summation_n[(name.replace('預', ''), token)] += int(n) if n != '?' else math.nan
        return val

    def verif_count_n(name, val, ws, idx, **kwargs):
        for l in val.splitlines():
            token, n, d = re.search(r'([\w\W]+)' + pattern_detectfrac, l).groups()
            if d == '?' or summation_n[(name.replace('預', ''), token)] != int(d):
                print('BUG @ ', name, l)
        return val

    sanitize('I', sum_count_n)
    sanitize('I', verif_count_n)

    unique_denom = {}
    numrtr_left = {}

    def correct(name, val, ws, idx, **kwargs):
        showname = name.replace('預', '')
        totused = ws['G'][idx].value
        for l in val.splitlines():
            token, n, d = re.search(r'([\w\W]+)' + pattern_detectfrac, l).groups()
            assert d != '?'
            assert n != '?'
            n, d = map(int, (n, d))
            assert showname, token not in unique_denom or unique_denom[showname, token] == d
            if (showname, token) not in numrtr_left: # first time
                numrtr_left[showname, token] = d - n
            else:
                numrtr_left[showname, token] -= n
            totused -= n
        assert totused == 0
        return val

    sanitize('I', correct)
    for v in numrtr_left.values(): assert v == 0


def has_unkown(name, val, ws, idx, **kwargs):
    if '?' in val:
        print(name)
        print(val)
    return val


def new_year2024(name, val, ws, idx, **kwargs):
    return (val
            .replace('113/2/7 全日八(6/8)', '113/2/7 特殊全日八(6/6)')
            .replace('113/2/7 晚四(2/4)', '113/2/7 特殊晚四(2/2)')
            .replace('113/2/7 全日八(2/8)', '113/2/7 特殊全日八(2/6)')
            )


if __name__ == '__main__':
    sanitize_header()
    sanitize_br()
    sanitize('I', hasparanthesis, doprint=False)
    sanitize('I', compress, doprint=False)
    precheck_confirm()
    sanitize('I', iter_solve)
    sanitize('I', new_year2024)
    verify_count()
    sanitize('I', has_unkown)
    wb.save(target)
