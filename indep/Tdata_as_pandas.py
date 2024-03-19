"""xlsx表格 轉換成pandas資料結構 存入二進位檔.pickle"""
import os
import re
import openpyxl as ox
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import numpy as np
import pandas as pd

full_source = 'Tdata_compiled\替代役訓練及管理中心役男退役管制名冊(持續更新)0312 253T.xlsx'

workbook: Workbook = ox.load_workbook(full_source)
worksheet: Worksheet = workbook['113.02.01-業務助理編號從8901開始']
tz = [int(x.value) for x in np.array(worksheet[f'A3:A{worksheet.max_row}'])[:, 0]]
namez = [x.value for x in np.array(worksheet[f'D3:D{worksheet.max_row}'])[:, 0]]
uidz = [int(x.value) for x in np.array(worksheet[f'M3:M{worksheet.max_row}'])[:, 0]]
unit = [x.value for x in np.array(worksheet[f'C3:C{worksheet.max_row}'])[:, 0]]
full = pd.DataFrame({'uid': uidz, 'name': namez, 'tnum': tz, 'unit': unit})
full = full.sort_values(by='uid').reset_index(drop=True)

start = max(uidz)
tz_treated_till = max(tz)
for fn in sorted(os.listdir('Tdata_compiled')):
    m = re.match('database(\d+).xlsx', fn)
    if m and int(m.group(1)) > tz_treated_till:
        read = pd.read_excel(os.path.join('Tdata_compiled', fn))
        read.to_pickle(os.path.join('Tdata_compiled', fn.replace('xlsx', 'pickle')))

        read = read.sort_values(by=['梯次', '預退日期', '出生日期', '名']).reset_index()
        read.insert(0, 'uid', np.arange(read.shape[0]) + start)

        for _bad_labels in read.columns:
            if _bad_labels not in ['uid', '名', '梯次']:
                read.drop(_bad_labels)

        read.rename(columns={"名": "name", "梯次": "tnum"})
        read.insert(3, "unit", ["未知" for x in range((read.shape[0]))])

        full: pd.DataFrame = read if read is None else pd.concat([full, read])

full.to_pickle(os.path.join('Tdata_compiled', 'allT.pickle'))
