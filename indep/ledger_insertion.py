"""從梯次的.pickle 檔內，新增 役男榮譽假清冊 的分頁"""
import collections
import math
import operator
import re
import shutil
from collections import defaultdict

import pandas as pd
import openpyxl as ox
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

pattern_date3 = '(\d{3})(\d{2})(\d{2})'
pattern_usage = '(\([\d/?]+\))'
pattern_detectfrac = '\(([\d?]+)/([\d?]+)\)'

xls_file = 'holiday_ledger/組改期間役男榮譽假清冊3_改.xlsx'
out_file = 'holiday_ledger/組改期間役男榮譽假清冊3_改temp.xlsx'


def inflate_before_northen_office(ordered_names):
    wb: Workbook = ox.load_workbook(xls_file)
    sheet_names = set(wb.sheetnames)
    sheet = wb[wb.sheetnames[-1]]
    for _r in sheet.rows: break
    header = [c.value for c in _r]
    for name in ordered_names:
        if name not in sheet_names:
            print(f'creating {name}')
            wb.create_sheet(name, len(wb.sheetnames) - 1)
            sheet = wb[name]
            sheet.append(header)
    wb.save(out_file)


if __name__ == '__main__':
    data253 = pd.read_pickle('Tdata_compiled/database253.pickle')
    inflate_before_northen_office(data253.sort_values(by=['預退日期']).名)
