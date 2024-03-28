import collections

from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import numpy as np
from openpyxl import Workbook
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.worksheet import Worksheet
import openpyxl as ox
import re
from datetime import date as Date, datetime as Datetime, timedelta as TimeDelta, time as Time
from enum import Enum
import compile_monthly_ledger as uml
from interactions import interaction, start_interaction
from compile_monthly_ledger import workhours, grab_till
# import prettytable as pt
from tabulate import tabulate
import sys
import weekly.weekly as wk

import pandas as pd

# target = r'D:\東區督考科業務\1.東區業務\組改期間榮譽假系統\組改期間役男榮譽假清冊自動化.xlsx'  # TODO
target = 'holiday_ledger/組改期間役男榮譽假清冊3_改.xlsx'  # TODO
wb_global: Worksheet = None

full_data = pd.read_pickle('Tdata_compiled/allT.pickle')

pattern_detectfrac = r'\(([\d?]+)/([\d?]+)\)'
pattern_detectparan = r'\(([\d?]+)\)'
pattern_date3 = r'(\d{3})/(\d+)/(\d+)'


def rows_no_header(ws):
    header = None
    for row in ws.rows:
        if header is None:
            header = row
            continue
        yield tuple(c for c in row)


def id2name(uid):
    uid = int(uid)
    # res: pd.DataFrame = full_data[full_data.uid.isin([uid-8000, uid, 8000 + uid])]
    res: pd.DataFrame = full_data[full_data.uid == uid % 8000]
    assert len(res) == 1
    return res.name.values[0]


def name2unit(name):
    # res: pd.DataFrame = full_data[full_data.uid.isin([uid-8000, uid, 8000 + uid])]
    res: pd.DataFrame = full_data[full_data.name == name]
    assert len(res) == 1
    return res.unit.values[0]


def obtain_name(uid_or_name):
    if isinstance(uid_or_name, int):
        uid = int(uid_or_name)
        name = id2name(uid)
    elif re.match(r'\d{3}\d?', uid_or_name):
        uid = int(uid_or_name)
        name = id2name(uid)
    else:
        name = uid_or_name
    return name


def key(x):
    exp = tuple(int(xx) for xx in re.findall(r'\d+', x[7])) if x[7] is not None else (111, 1, 1)
    firstuse = x[8].splitlines()[0].replace('*預', '') if x[8] is not None else '120/1/1'
    lastuse = x[8].splitlines()[-1].replace('*預', '') if x[8] is not None else '120/1/1'
    # lastline = re.match(pattern_date3, lastline).groups()
    firstuse = tuple(-int(xx) for xx in re.match(pattern_date3, firstuse).groups())
    lastuse = tuple(-int(xx) for xx in re.match(pattern_date3, lastuse).groups())
    comp = int(x[5]) - int(x[6]) == 0, exp, x[0] is not None, int(x[5]) - int(x[6])
    if int(x[5]) - int(x[6]) == 0:
        cmp = 1, lastuse, '*預' not in x[8], x[0] is not None, firstuse, exp
        # 無空後, 快過期者先, 過期日期，預價先
    else:
        cmp = 0, lastuse, firstuse, exp
        # 有空先，早過期者先，則快過期者先
    return cmp


def sort_toppriority(ws):
    rvals = []
    for row in rows_no_header(ws):
        rvals.append(tuple(x.value for x in row))

    rvals.sort(key=key)

    for row, rowv in zip(rows_no_header(ws), rvals):
        for cell, val in zip(row, rowv):
            cell.value = val


@interaction
def use_holiday(uid_or_name, timestring: str):
    """

    :param uid_or_name: 123 or 8123 or 王大明
    :param timestring: 1130325 0830-1130326 1330
    :return:
    """
    timestring = timestring.replace('-', '~')
    timestring = re.sub(r'(1\d{2})(\d{2})(\d{2})', r'\g<1>/\g<2>/\g<3>', timestring)

    name = obtain_name(uid_or_name)

    if wb_global is None: load()
    ws: Worksheet = wb_global[name]

    fr, to = grab_till(timestring)
    yr, mn, dy, frtime = fr
    yr2, mn2, dy2, totime = to
    usage = sum(
        Datetime.combine(Date(yr + 1911, mn, dy), frtime) <=
        x <=
        Datetime.combine(Date(yr2 + 1911, mn2, dy2), totime) for x in uml.workhours
    )
    decompress = (f'{yr}/{mn}/{dy} {frtime.hour:02}{frtime.minute:02}'
                  f'~{yr2}/{mn2}/{dy2} {totime.hour:02}{totime.minute:02} 自動特')

    denom = usage
    while usage != 0:
        sort_toppriority(ws)
        if ws['F2'].value - ws['G2'].value == 0:
            ws.append((None,) * 3 + ('預假', None, usage, usage, ws['H2'].value, decompress + f'({usage}/{denom})'))
            print(f'{name} 的 "{decompress}" 需要預假 {usage}/{denom}')
            sort_toppriority(ws)
            break
        use = min(ws['F2'].value - ws['G2'].value, usage)
        usage = usage - use
        line = f'{decompress}({use}/{denom})'
        print(f'{name} {ws["D2"].value}(~{ws["H2"].value};{ws["J2"].value}) "" 添加 "{line}"')
        if ws['I2'].value is None: ws['I2'].value = ''
        ws['I2'].value = '\n'.join(ws['I2'].value.splitlines() + [line])
        ws['G2'].value += use
    refresh_format(name)
    sort_toppriority(ws)
    print_sheet(name)


@interaction
def add_holiday_quota_auto3month(uid_or_name, hours: int, date_added: str, reason: str):
    _ = re.match(r'(\d{3})-?(\d{2})-?(\d{2})', date_added)
    y, m, d = _.groups()
    expiration = f'{y}-{int(m) + 3:02}-{int(d):02}'
    add_holiday_quota(uid_or_name, hours, date_added, expiration, reason)


@interaction
def add_holiday_quota(uid_or_name, hours: int, period: str, expiration: str, reason: str):
    name = obtain_name(uid_or_name)
    if not isinstance(hours, int): hours = int(hours)
    _ = re.match(r'(\d{3})-?(\d{2})-?(\d{2})', expiration)
    _ = _.groups()
    expiration = '{}-{:02}-{:02}'.format(*(int(x) for x in _))

    if wb_global is None: load()
    ws: Worksheet = wb_global[name]
    preempt = collections.defaultdict(int)
    denom = collections.defaultdict(int)

    for r in rows_no_header(ws):
        is_not_prempt = r[0].value
        if is_not_prempt:
            usage = r[8].value if r[8].value is not None else ''
            for l in usage.splitlines():
                token, n, d = re.search(r'([\w\W]+)' + pattern_detectfrac, l).groups()
                if token.startswith('*預'): preempt[token] -= int(n)
        else:
            usage = r[8].value if r[8].value is not None else ''
            assert '\n' not in usage  # must be one line!!!
            l = usage
            token, n, d = re.search(r'([\w\W]+)' + pattern_detectfrac, l).groups()
            preempt['*預' + token] += int(n)
            denom['*預' + token] = int(n)
    if all(v == 0 for v in preempt.values()):
        print('沒有預假需要核銷~ 新增空白假')
        ws.append((name2unit(name), '替代役男', name, '榮譽假', period, hours, 0, expiration, None, reason))
    else:
        for k, v in preempt.items():
            if v > 0:
                used = min(hours, v)
                usage = k + f'({used}/{denom[k]})'
                ws.append(
                    _ := (name2unit(name), '替代役男', name, '榮譽假', period, hours, used, expiration, usage, reason))
                break
        print('!!有預假自動核銷' + ';'.join(f"{x}" for x in _))
    sort_toppriority(ws)
    print_sheet(name)
    for r in rows_no_header(ws):
        print(key([x.value for x in r]))
    print()


binbydate = collections.defaultdict(set)  #
holidaytypes = collections.defaultdict(set)
workdays = uml.gen_workdays()


def _compile_date(date: Date):
    date = Datetime.strptime(date.strftime('%Y-%m-%d'), '%Y-%m-%d')
    imminent_discharge = full_data[np.abs(full_data.discharge - date) < TimeDelta(days=2)]
    if len(imminent_discharge):
        print('IMMINENT DISCHARGE 即將退役!')
        print(imminent_discharge)
    remaining = full_data[full_data.discharge > Datetime.strptime(date.strftime('%Y-%m-%d'), '%Y-%m-%d')]

    for name in filter(lambda x: x not in ['胡力元', '廖育佐'], remaining.name):
        if wb_global is None: load()
        ws: Worksheet = wb_global[name]

        for cell, holidaytype in zip(ws['I'][1:], ws['D'][1:]):  # I是使用紀錄，1是跳掉開頭的"使用紀錄"
            if cell.value is None: continue
            content: str = cell.value
            for li, line in enumerate(content.splitlines()):  # each line is one (partial usage of holiday)
                if line.startswith('*預'): continue  # already used, we skip
                _ = re.search(r'([\w\W]+)' + pattern_detectfrac, line)
                if _ is None:
                    print(f'unable to match!!: {name}{line}, trying simple paranthesis')
                    _ = re.search(r'([\w\W]+)' + pattern_detectparan, line)
                    token, n = _.groups()
                    _lines = content.splitlines()
                    _lines[li] += f'({n}/{n})'
                    print(f'auto appending fraction ({n}/{n}) and repeating PLEASE CHECK THE FOLLOWING!!')
                    print(_lines[li])
                    content = '\n'.join(_lines)
                    line = content.splitlines()[li]
                    _ = re.search(r'([\w\W]+)' + pattern_detectfrac, line)

                token, n, d = _.groups()

                _fr, _to = grab_till(token)
                yr, mn, dy, starttime = _fr
                yr2, mn2, dy2, endtime = _to
                yr = int(yr) + 1911
                yr2 = int(yr2) + 1911
                yr, mn, dy, yr2, mn2, dy2 = map(int, [yr, mn, dy, yr2, mn2, dy2])

                _fr = Datetime.combine(Date(year=yr, month=mn, day=dy), starttime)
                _to = Datetime.combine(Date(year=yr2, month=mn2, day=dy2), endtime)

                dayz = []
                d_ = Date(year=yr, month=mn, day=dy)
                d_to = Date(year=yr2, month=mn2, day=dy2)
                while d_ <= d_to:
                    if d_ in workdays:
                        dayz.append(d_)
                    d_ += TimeDelta(days=1)

                if holidaytype.value is None:
                    holidaytypev = '預'
                elif '預' in holidaytype.value:
                    holidaytypev = '預'
                elif '榮' in holidaytype.value:
                    holidaytypev = '榮'
                elif '補' in holidaytype.value:
                    holidaytypev = '補'
                elif '公' in holidaytype.value:
                    holidaytypev = '公'
                else:
                    print(holidaytype.value)
                    raise ValueError

                for day in dayz:
                    starttime = Time(17, 30)
                    endtime = Time(8, 30)
                    for t in wk.allday:
                        if _fr < Datetime.combine(day, t) < _to:
                            starttime = min(t, starttime)
                            endtime = max(t, endtime)
                    starttime = (Datetime.combine(day, starttime) - TimeDelta(minutes=15)).time()
                    endtime = (Datetime.combine(day, endtime) + TimeDelta(minutes=45)).time()

                    iso_signature = f'{starttime.hour:02}:{starttime.minute:02}-{endtime.hour:02}:{endtime.minute:02}'

                    binbydate[day].add((name, iso_signature))
                    holidaytypes[(day, name, iso_signature)].add(holidaytypev)
                    # print((day, name, signature, holidaytypev))


@interaction
def compile_weekdays(datestr: str):
    try:
        yr, mn, dy, yr2, mn2, dy2 = re.match(r'(\d{3})-?(\d{2})-?(\d{2})[\-~]](\d{3})-?(\d{2})-?(\d{2})',
                                             datestr).groups()
        yr, mn, dy = map(int, [yr, mn, dy])
        yr2, mn2, dy2 = map(int, [yr2, mn2, dy2])
        if yr < 1911: yr += 1911
        if yr2 < 1911: yr2 += 1911
        start = Date(yr, mn, dy)
        end = Date(yr2, mn2, dy2)
        while start <= end:
            compile_weekday(f'{start.year - 1911}{start.month:02}{start.day:02}')
            start += TimeDelta(days=1)

    except:
        compile_weekday(datestr)


def compile_weekday(datestr: str):
    yr, mn, dy = re.match(r'(\d{3})-?(\d{2})-?(\d{2})', datestr).groups()
    yr = int(yr)
    if yr < 1911: yr += 1911
    yr, mn, dy = map(int, [yr, mn, dy])
    date = Date(yr, mn, dy)

    if len(binbydate) == 0:
        _compile_date(date)
        print('compiled binbydate')

    absence = binbydate[date]
    with wk.prepare(date) as f:
        for name, iso_signature in absence:
            holiday_type = '+'.join(holidaytypes[(date, name, iso_signature)])
            f.add(name, full_data[full_data.name == name].dorm.iloc[0], holiday_type, iso_signature)

    date.weekday()
    print('END of compile_weekday')


@interaction
def refresh_format(uid_or_name):
    name = obtain_name(uid_or_name)

    if wb_global is None: load()
    ws: Worksheet = wb_global[name]

    for r in rows_no_header(ws):
        for c in r:
            c.font = Font()
            c.border = Border()
            c.alignment = Alignment()


@interaction
def save_new(postfix):
    wb_global.save(target.replace('.xlx', postfix + '.xlsx'))


@interaction
def save():
    wb_global.save(target)


@interaction
def load():
    global wb_global
    wb_global = ox.load_workbook(target)


@interaction
def print_sheet(uid_or_name):
    name = obtain_name(uid_or_name)
    if wb_global is None: load()
    ws = wb_global[name]
    rows = []
    for r in rows_no_header(ws):
        rows.append([c.value for ic, c in enumerate(r) if ic < 10])

    print(tabulate(rows, headers=[ws[f'{c}1'].value for c in 'ABCDEFGHIJ'], tablefmt='fancy_grid'))


@interaction
def calc_remainder(uid_or_name):
    name = obtain_name(uid_or_name)
    if wb_global is None: load()
    ws = wb_global[name]
    rows = []
    left = 0
    for r in rows_no_header(ws):
        rows.append([c.value for c in r])
        left += (r[5].value - r[6].value)
    print(f'{name} has {left} hrs left')


if __name__ == '__main__':
    if len(sys.argv) > 1:
        with open(sys.argv[1], 'r') as f:
            lines = f.readlines()


        def mygen():
            for l in lines: yield l.strip()


        g = mygen()


        def myinput(prompt=''):
            try:
                n = next(g)
                print(prompt, end=f'{n}\n')
                return n
            except StopIteration:
                return input(prompt)


        start_interaction(myinput)
    else:
        start_interaction()
    # id2name(519)
    # id2name(8519)
    # use_holiday(519, '1130315 0900-19 1700')
    # append_new_holiday(519, '1130315 0900-19 1700(8)')
    # refresh_format(519)
    # print()
